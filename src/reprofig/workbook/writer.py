"""Readable, deterministic Excel projection of a publication dataset."""

from __future__ import annotations

import csv
import io
import json
import os
import tempfile
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence

from ..schema import deterministic_json
from ..tables import safe_filename_token
from .models import PublicationDataset, PublicationTable
from .statistics import NORMALIZED_COLUMNS, statistics_rows

BASE_SHEETS = (
    "README", "Figures", "Data_Index", "Statistics", "Test_Families",
    "Verification", "Dictionary",
)
EXCEL_MAX_ROWS = 1_048_576
EXCEL_MAX_COLUMNS = 16_384
EXCEL_MAX_CELL_CHARS = 32_767


@dataclass(frozen=True)
class WorkbookRenderResult:
    path: Path
    sheet_map: dict[str, str]
    logical_fingerprint: str


def _openpyxl() -> Any:
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover - dependency-specific
        raise RuntimeError(
            "publication workbooks require the Excel extra: pip install reprofig[excel]"
        ) from exc
    return openpyxl


def data_sheet_map(dataset: PublicationDataset) -> dict[str, str]:
    result: dict[str, str] = {}
    used = set(BASE_SHEETS)
    for index, table in enumerate(dataset.tables, start=1):
        prefix = f"D{index:03d}_"
        slug = safe_filename_token(table.name, fallback="data")[: 31 - len(prefix)]
        candidate = (prefix + slug)[:31]
        suffix = 1
        while candidate.casefold() in {value.casefold() for value in used}:
            tail = f"_{suffix}"
            candidate = (prefix + slug[: 31 - len(prefix) - len(tail)] + tail)[:31]
            suffix += 1
        used.add(candidate)
        result[str(table.table_id)] = candidate
    return result


def _safe_cell(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, (dict, list, tuple, set)):
        value = deterministic_json(value)
    if isinstance(value, str) and len(value) > EXCEL_MAX_CELL_CHARS:
        raise ValueError(f"cell text exceeds Excel's {EXCEL_MAX_CELL_CHARS}-character limit")
    return value


def _append(ws: Any, values: Sequence[Any]) -> None:
    if len(values) > EXCEL_MAX_COLUMNS:
        raise ValueError(f"worksheet {ws.title!r} exceeds Excel's column limit")
    if ws.max_row >= EXCEL_MAX_ROWS:
        raise ValueError(f"worksheet {ws.title!r} exceeds Excel's row limit")
    ws.append([_safe_cell(value) for value in values])
    for cell in ws[ws.max_row]:
        if isinstance(cell.value, str) and cell.value.startswith(("=", "+", "-", "@")):
            cell.data_type = "s"


def _write_rows(ws: Any, headers: Sequence[str], rows: Iterable[Mapping[str, Any]]) -> None:
    _append(ws, list(headers))
    for row in rows:
        _append(ws, [row.get(header) for header in headers])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for column in ws.columns:
        letter = column[0].column_letter
        width = max((len(str(cell.value or "")) for cell in column[:200]), default=0) + 2
        ws.column_dimensions[letter].width = min(max(width, 10), 50)


def _read_csv(table: PublicationTable) -> list[list[str]]:
    if table.contents is None:
        return []
    rows = list(csv.reader(io.StringIO(table.contents, newline="")))
    if len(rows) > EXCEL_MAX_ROWS:
        raise ValueError(f"table {table.table_id} exceeds Excel's row limit")
    if any(len(row) > EXCEL_MAX_COLUMNS for row in rows):
        raise ValueError(f"table {table.table_id} exceeds Excel's column limit")
    for row_index, row in enumerate(rows, start=1):
        for column_index, value in enumerate(row, start=1):
            if len(value) > EXCEL_MAX_CELL_CHARS:
                raise ValueError(
                    f"table {table.table_id} cell R{row_index}C{column_index} exceeds Excel's cell limit"
                )
    return rows


def workbook_projection(dataset: PublicationDataset, sheet_map: Mapping[str, str]) -> dict[str, list[list[Any]]]:
    """Return the canonical visible-cell projection used by writer and validator."""

    figures_headers = (
        "figure_id", "figure_record_sha256", "schema", "profile", "title",
        "original_stem", "source_label", "display_order",
    )
    figures = [[getattr(item, name, None) for name in figures_headers] for item in dataset.figures]
    index_headers = (
        "table_id", "sheet", "sha256", "name", "purpose", "row_count", "column_count",
        "figure_id", "table_name", "table_index", "figure_record_sha256",
    )
    index_rows: list[list[Any]] = []
    for table in dataset.tables:
        uses = table.occurrences or [None]
        for occurrence in uses:
            index_rows.append([
                table.table_id, sheet_map.get(str(table.table_id)), table.sha256, table.name,
                table.purpose, table.row_count, table.column_count,
                occurrence.figure_id if occurrence else None,
                occurrence.table_name if occurrence else None,
                occurrence.table_index if occurrence else None,
                occurrence.figure_record_sha256 if occurrence else None,
            ])
    stat_rows = statistics_rows(dataset)
    statistics = [[row.get(name) for name in NORMALIZED_COLUMNS] for row in stat_rows]
    family_headers = ("family_id", "label", "method", "test_ids", "display_order")
    families = [[
        item.family_id, item.label, item.method, deterministic_json(item.test_ids), item.display_order
    ] for item in dataset.test_families]
    verification_headers = (
        "verification_id", "subject_id", "status", "check", "message", "details",
    )
    verification = [[
        item.verification_id, item.subject_id, item.status, item.check, item.message,
        deterministic_json(item.details),
    ] for item in dataset.verification]
    projection: dict[str, list[list[Any]]] = {
        "Figures": [list(figures_headers), *figures],
        "Data_Index": [list(index_headers), *index_rows],
        "Statistics": [list(NORMALIZED_COLUMNS), *statistics],
        "Test_Families": [list(family_headers), *families],
        "Verification": [list(verification_headers), *verification],
    }
    for table in dataset.tables:
        if table.contents is not None:
            projection[sheet_map[str(table.table_id)]] = _read_csv(table)
    return projection


def render_workbook(
    dataset: PublicationDataset,
    output_path: str | os.PathLike[str],
    *,
    overwrite: bool = False,
) -> WorkbookRenderResult:
    """Render a fresh formula-free workbook and atomically promote it."""

    errors = dataset.validate()
    if errors:
        raise ValueError("invalid publication dataset:\n- " + "\n- ".join(errors))
    target = Path(output_path)
    if target.suffix.lower() != ".xlsx":
        raise ValueError("publication workbook output must use .xlsx")
    if target.exists() and not overwrite:
        raise FileExistsError(target)
    target.parent.mkdir(parents=True, exist_ok=True)
    sheet_map = data_sheet_map(dataset)
    projection = workbook_projection(dataset, sheet_map)
    openpyxl = _openpyxl()
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)
    fixed = datetime(2000, 1, 1, tzinfo=timezone.utc)
    workbook.properties.creator = "ReproFig"
    workbook.properties.title = "Publication source data"
    workbook.properties.created = fixed
    workbook.properties.modified = fixed

    readme = workbook.create_sheet("README")
    readme_rows = [
        ("Publication ID", dataset.publication_id),
        ("Profile", dataset.profile),
        ("Statistics coverage", dataset.statistics_coverage),
        ("Logical evidence SHA-256", dataset.fingerprint()),
        ("Canonical evidence", "Embedded ReproFig record; worksheet cells are readable projections."),
        ("Coverage warning", "analysis_complete is a declaration, not proof that undisclosed analyses never occurred."),
    ]
    for row in readme_rows:
        _append(readme, row)
    readme.column_dimensions["A"].width = 28
    readme.column_dimensions["B"].width = 90
    readme.freeze_panes = "A2"

    for name in ("Figures", "Data_Index", "Statistics", "Test_Families", "Verification"):
        ws = workbook.create_sheet(name)
        rows = projection[name]
        _write_rows(ws, rows[0], [dict(zip(rows[0], row)) for row in rows[1:]])

    dictionary = workbook.create_sheet("Dictionary")
    dictionary_rows = [
        {"field": "*_exact", "meaning": "Authoritative exact text from the source record."},
        {"field": "*_numeric", "meaning": "Convenience numeric projection; exact text remains authoritative."},
        {"field": "statistics_coverage", "meaning": "incomplete, figure_complete, analysis_complete, or not_applicable."},
        {"field": "table_id", "meaning": "Stable identity derived from exact canonical CSV bytes."},
        {"field": "figure_record_sha256", "meaning": "SHA-256 fingerprint of the canonical source figure record."},
    ]
    _write_rows(dictionary, ("field", "meaning"), dictionary_rows)

    for table in dataset.tables:
        if table.contents is None:
            continue
        ws = workbook.create_sheet(sheet_map[str(table.table_id)])
        rows = projection[ws.title]
        for row in rows:
            _append(ws, row)
        if rows:
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions

    descriptor = deterministic_json({
        "publication_id": dataset.publication_id,
        "fingerprint": dataset.fingerprint(),
        "sheet_map": sheet_map,
    })
    workbook.custom_doc_props.append(
        openpyxl.packaging.custom.StringProperty(name="ReproFig", value=descriptor)
    )

    handle, candidate_name = tempfile.mkstemp(prefix=f".{target.stem}-", suffix=".xlsx", dir=target.parent)
    os.close(handle)
    candidate = Path(candidate_name)
    try:
        workbook.save(candidate)
        reopened = openpyxl.load_workbook(candidate, read_only=True, data_only=False)
        reopened.close()
        os.replace(candidate, target)
    finally:
        try:
            candidate.unlink()
        except OSError:
            pass
    return WorkbookRenderResult(target, sheet_map, dataset.fingerprint())


__all__ = [
    "BASE_SHEETS",
    "WorkbookRenderResult",
    "data_sheet_map",
    "render_workbook",
    "workbook_projection",
]
