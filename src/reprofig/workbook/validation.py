"""Cross-check Excel worksheet projections against embedded canonical evidence."""

from __future__ import annotations

import os
import math
import zipfile
from pathlib import Path
from typing import Any, Mapping, Sequence

from ..artifacts import extract_record
from ..schema import deterministic_json
from ..validation import ValidationReport, privacy_leaks, validate_record
from .evidence import publication_dataset_from_record
from .writer import BASE_SHEETS, workbook_projection


def _openpyxl() -> Any:
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError(
            "publication workbook validation requires: pip install reprofig[excel]"
        ) from exc
    return openpyxl


def _cell_value(value: Any) -> Any:
    if isinstance(value, (dict, list, tuple, set)):
        return deterministic_json(value)
    return value


def _sheet_values(ws: Any) -> list[list[Any]]:
    rows = [[cell.value for cell in row] for row in ws.iter_rows()]
    while rows and all(value is None for value in rows[-1]):
        rows.pop()
    width = max((len(row) for row in rows), default=0)
    while width and all((len(row) < width or row[width - 1] is None) for row in rows):
        width -= 1
    return [row[:width] for row in rows]


def _compare_sheet(
    report: ValidationReport,
    workbook: Any,
    name: str,
    expected: Sequence[Sequence[Any]],
) -> None:
    if name not in workbook.sheetnames:
        report.add("error", "workbook_sheet_missing", f"Expected worksheet {name!r} is missing", name)
        return
    actual = _sheet_values(workbook[name])
    normalized_expected = [[_cell_value(value) for value in row] for row in expected]
    if actual == normalized_expected:
        report.checks.append(f"sheet:{name}")
        return
    rows = max(len(actual), len(normalized_expected))
    for row_index in range(rows):
        actual_row = actual[row_index] if row_index < len(actual) else []
        expected_row = normalized_expected[row_index] if row_index < len(normalized_expected) else []
        columns = max(len(actual_row), len(expected_row))
        for column_index in range(columns):
            actual_value = actual_row[column_index] if column_index < len(actual_row) else None
            expected_value = expected_row[column_index] if column_index < len(expected_row) else None
            equal = actual_value == expected_value
            if (
                not equal
                and isinstance(actual_value, (int, float))
                and isinstance(expected_value, (int, float))
                and not isinstance(actual_value, bool)
                and not isinstance(expected_value, bool)
            ):
                equal = math.isclose(
                    float(actual_value), float(expected_value), rel_tol=1e-12, abs_tol=1e-15
                )
            if not equal:
                report.add(
                    "error",
                    "workbook_cell_mismatch",
                    f"Expected {expected_value!r}, found {actual_value!r}",
                    f"{name}!R{row_index + 1}C{column_index + 1}",
                )
                return


def _scan_office_package(report: ValidationReport, path: Path) -> None:
    try:
        with zipfile.ZipFile(path) as archive:
            if len(archive.namelist()) != len(set(archive.namelist())):
                report.add("error", "duplicate_zip_member", "Office package has duplicate members")
                return
            for info in archive.infolist():
                if info.file_size > 100 * 1024 * 1024:
                    report.add("error", "office_member_too_large", "Office member exceeds safety limit", info.filename)
                    continue
                try:
                    text = archive.read(info).decode("utf-8")
                except (UnicodeDecodeError, RuntimeError, zipfile.BadZipFile):
                    continue
                for code, match in privacy_leaks(text):
                    report.add("error", f"private_{code}", f"Private value found: {match}", info.filename)
    except zipfile.BadZipFile as exc:
        report.add("error", "invalid_excel_package", str(exc))


def validate_publication_workbook(
    path: str | os.PathLike[str],
    *,
    require_complete: bool = True,
    public_safety: bool | None = None,
) -> ValidationReport:
    target = Path(path)
    report = ValidationReport(path=str(target))
    try:
        record = extract_record(target)
    except Exception as exc:
        report.add("error", "embedded_record_invalid", str(exc), "reprofig/manifest.json")
        return report
    report.profile = record.distribution_profile
    integrity = validate_record(record, require_complete=False)
    report.issues.extend(integrity.issues)
    report.checks.extend(f"record:{check}" for check in integrity.checks)
    if not integrity.valid:
        return report
    try:
        dataset = publication_dataset_from_record(record)
    except Exception as exc:
        report.add("error", "publication_evidence_invalid", str(exc), "publication_workbook")
        return report
    extension = record.extensions["publication_workbook"]
    if require_complete and dataset.profile == "master":
        protected = set(extension.get("protected_section_ids", [])) if isinstance(extension, Mapping) else set()
        missing = [
            str(table.table_id)
            for table in dataset.tables
            if table.contents is None and str(table.table_id) not in protected
        ]
        if missing:
            report.add("error", "publication_data_incomplete", "Master workbook lacks embedded tables: " + ", ".join(missing))

    sheet_map = extension.get("sheet_map") or {}
    expected_map = {str(table.table_id): sheet_map.get(str(table.table_id)) for table in dataset.tables}
    if any(not value for value in expected_map.values()):
        report.add("error", "sheet_mapping_incomplete", "Embedded worksheet mapping is incomplete")
        return report
    try:
        workbook = _openpyxl().load_workbook(target, read_only=True, data_only=False)
    except Exception as exc:
        report.add("error", "excel_open_failed", str(exc))
        return report
    try:
        expected_names = list(BASE_SHEETS) + [
            expected_map[str(table.table_id)] for table in dataset.tables if table.contents is not None
        ]
        if workbook.sheetnames != expected_names:
            report.add(
                "error", "workbook_sheet_order", f"Expected worksheets {expected_names!r}, found {workbook.sheetnames!r}"
            )
        projection = workbook_projection(dataset, expected_map)
        for name, rows in projection.items():
            _compare_sheet(report, workbook, name, rows)
        report.checks.extend(("publication_fingerprint", "worksheet_mapping", "embedded_evidence"))
    finally:
        workbook.close()

    if public_safety is None:
        public_safety = dataset.profile in {"public", "minimal_public"}
    if public_safety:
        _scan_office_package(report, target)
        report.checks.append("office_package_privacy")
    return report


__all__ = ["validate_publication_workbook"]
