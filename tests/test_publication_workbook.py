from __future__ import annotations

import zipfile

from openpyxl import load_workbook

from reprofig import build_publication_workbook
from reprofig.api import build_record
from reprofig.artifacts import embed_file, extract_record
from reprofig.workbook import publication_dataset_from_record, validate_publication_workbook
from reprofig.crypto.encryption import decrypt_sections


def _artifact(tmp_path):
    source = tmp_path / "plain.svg"
    source.write_text('<svg xmlns="http://www.w3.org/2000/svg"><path d="M0 0"/></svg>')
    record = build_record(
        title="Test figure",
        producer={"package": "tests", "version": "1"},
        plotted_data=[
            {"group": "control", "value": "=1+1"},
            {"group": "treated", "value": "2.0000000000000000001"},
        ],
        statistics=[
            {
                "test_id": "welch-1",
                "test_name": "Welch t-test",
                "n_a": 1,
                "n_b": 1,
                "statistic": "3.141592653589793238",
                "p": "0.01234567890123456789",
                "alternative": "two_sided",
            }
        ],
        column_classification={"group": "safe", "value": "safe"},
    )
    artifact = tmp_path / "figure.svg"
    embed_file(source, record, output_path=artifact)
    return artifact


def test_publication_workbook_roundtrip_and_tamper_detection(tmp_path):
    artifact = _artifact(tmp_path)
    output = tmp_path / "Publication-source-data.xlsx"
    result = build_publication_workbook(artifact, output)
    assert result.valid
    assert result.figure_count == 1
    assert result.unique_table_count == 1
    assert result.statistic_count == 1

    workbook = load_workbook(output, data_only=False)
    assert workbook.sheetnames[:7] == [
        "README", "Figures", "Data_Index", "Statistics", "Test_Families",
        "Verification", "Dictionary",
    ]
    data_sheet = workbook[workbook.sheetnames[7]]
    assert data_sheet["B2"].value == "=1+1"
    assert data_sheet["B2"].data_type == "s"
    assert workbook["Statistics"]["AA2"].value == "0.01234567890123456789"
    workbook.close()

    record = extract_record(output)
    dataset = publication_dataset_from_record(record)
    assert dataset.tables[0].contents.endswith("treated,2.0000000000000000001\n")
    assert validate_publication_workbook(output).valid

    tampered = tmp_path / "tampered.xlsx"
    with zipfile.ZipFile(output) as source, zipfile.ZipFile(tampered, "w") as destination:
        for info in source.infolist():
            contents = source.read(info)
            if info.filename.startswith("xl/worksheets/"):
                contents = contents.replace(b"2.0000000000000000001", b"999")
            destination.writestr(info, contents)
    output.write_bytes(tampered.read_bytes())
    report = validate_publication_workbook(output)
    assert not report.valid
    assert any(issue.code == "workbook_cell_mismatch" for issue in report.issues)


def test_publication_profiles_do_not_leak_removed_values(tmp_path):
    artifact = _artifact(tmp_path)
    minimal = tmp_path / "minimal.xlsx"
    result = build_publication_workbook(artifact, minimal, profile="minimal-public")
    assert result.valid
    workbook = load_workbook(minimal, read_only=True)
    assert len(workbook.sheetnames) == 7
    workbook.close()
    with zipfile.ZipFile(minimal) as archive:
        package = b"\n".join(archive.read(name) for name in archive.namelist())
    assert b"2.0000000000000000001" not in package


def test_publication_workbook_can_encrypt_a_source_table(tmp_path):
    artifact = _artifact(tmp_path)
    source_record = extract_record(artifact)
    table_id = f"table:{source_record.data_tables[0].sha256}"
    output = tmp_path / "protected.xlsx"
    result = build_publication_workbook(
        artifact,
        output,
        protected_sections=[table_id],
        encryption_password="authorized reader",
    )
    assert result.valid
    workbook = load_workbook(output, read_only=True)
    assert len(workbook.sheetnames) == 7
    workbook.close()
    aggregate = extract_record(output)
    decrypted = decrypt_sections(aggregate, password="authorized reader")
    assert "2.0000000000000000001" in str(decrypted[table_id])
