from __future__ import annotations

import json
import zipfile

from openpyxl import load_workbook
from PIL import Image
import pytest

from reprofig import build_publication_workbook
from reprofig.api import build_record
from reprofig.artifacts import embed_file, extract_record
from reprofig.crypto.keys import (
    generate_signing_key,
    load_signing_private_key,
    public_key_fingerprint,
    signing_public_bytes,
)
from reprofig.crypto.trust import TrustEntry, TrustStore
from reprofig.workbook import publication_dataset_from_record


def _records():
    shared = [
        {"participant": "participant-001", "group": "control", "value": "1.000000000000000001"},
        {"participant": "participant-002", "group": "treated", "value": "2.000000000000000002"},
    ]
    classification = {"participant": "private", "group": "safe", "value": "safe"}
    first = build_record(
        title="Figure A",
        plotted_data=shared,
        statistics=[{
            "test_id": "plot-a", "test_name": "Welch t-test",
            "p": "0.01234567890123456789", "statistic": "2.75",
            "n_a": 1, "n_b": 1, "alternative": "two_sided",
            "ci_low": "0.10000000000000001", "ci_high": "1.90000000000000009",
        }],
        column_classification=classification,
    )
    second = build_record(
        title="Figure B",
        plotted_data=shared,
        statistics=[{
            "test_id": "plot-b", "test_name": "Mann-Whitney U",
            "p": "0.03125000000000000000", "statistic": "1",
            "n_a": 1, "n_b": 1, "alternative": "two_sided",
            "p_adjusted": "0.06250000000000000000",
            "correction": "Holm", "family_id": "primary-comparisons",
        }],
        column_classification=classification,
    )
    third = build_record(
        title="Figure C",
        plotted_data=[{"participant": "participant-003", "group": "control", "value": 4}],
        statistics_status="not_applicable",
        column_classification=classification,
    )
    return first, second, third


def _mixed_artifacts(tmp_path):
    first, second, third = _records()
    folder = tmp_path / "figures"
    folder.mkdir()

    svg_raw = folder / "a-raw.svg"
    svg_raw.write_text('<svg xmlns="http://www.w3.org/2000/svg"><path d="M0 0"/></svg>')
    embed_file(svg_raw, first, output_path=folder / "a.svg")
    svg_raw.unlink()
    png_raw = folder / "a-raw.png"
    Image.new("RGB", (5, 5), "red").save(png_raw)
    embed_file(png_raw, first, output_path=folder / "a-copy.png")
    png_raw.unlink()

    import pikepdf

    pdf_raw = folder / "b-raw.pdf"
    with pikepdf.Pdf.new() as pdf:
        pdf.add_blank_page(page_size=(100, 100))
        pdf.save(pdf_raw)
    embed_file(pdf_raw, second, output_path=folder / "b.pdf")
    pdf_raw.unlink()
    xlsx_raw = folder / "b-raw.xlsx"
    with zipfile.ZipFile(xlsx_raw, "w") as archive:
        archive.writestr("xl/workbook.xml", "<workbook/>")
        archive.writestr(
            "[Content_Types].xml",
            '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types"/>',
        )
    embed_file(xlsx_raw, second, output_path=folder / "b-copy.xlsx")
    xlsx_raw.unlink()

    zip_raw = folder / "c-raw.zip"
    with zipfile.ZipFile(zip_raw, "w") as archive:
        archive.writestr("README.txt", "publication fixture")
    embed_file(zip_raw, third, output_path=folder / "c.zip")
    zip_raw.unlink()
    return folder, first, second, third


def _ledger(tmp_path):
    path = tmp_path / "all-tests.json"
    path.write_text(json.dumps({
        "schema": "reprofig-statistics-ledger/1",
        "analysis_id": "experiment-1",
        "coverage": "analysis_complete",
        "families": [{
            "family_id": "primary-comparisons", "label": "Primary comparisons",
            "method": "Holm", "test_ids": ["plot-b", "unplotted-qc"],
        }],
        "statistics": [
            {"test_id": "plot-a", "test_name": "Welch t-test", "p": "0.01234567890123456789", "statistic": "2.75", "n_a": 1, "n_b": 1, "alternative": "two_sided", "ci_low": "0.10000000000000001", "ci_high": "1.90000000000000009"},
            {"test_id": "plot-b", "test_name": "Mann-Whitney U", "p": "0.03125000000000000000", "statistic": "1", "n_a": 1, "n_b": 1, "alternative": "two_sided", "p_adjusted": "0.06250000000000000000", "correction": "Holm", "family_id": "primary-comparisons"},
            {"test_id": "unplotted-qc", "test_name": "Shapiro-Wilk", "p": "0.42345678901234567890", "statistic": "0.98", "n": 3, "displayed": False, "family_id": "primary-comparisons"},
        ],
    }), encoding="utf-8")
    return path


def test_mixed_carrier_analysis_complete_workbook_and_privacy_profiles(tmp_path):
    folder, first, _second, _third = _mixed_artifacts(tmp_path)
    ledger = _ledger(tmp_path)
    master_path = tmp_path / "master.xlsx"
    master = build_publication_workbook(
        folder, master_path, experiment_statistics=ledger
    )
    assert master.valid
    assert master.figure_count == 3
    assert master.unique_table_count == 2
    assert master.statistic_count == 3
    assert master.statistics_coverage == "analysis_complete"

    dataset = publication_dataset_from_record(extract_record(master_path))
    assert {item.test_id for item in dataset.statistics} == {
        "plot-a", "plot-b", "unplotted-qc"
    }
    assert sorted(table.contents for table in dataset.tables) == sorted({
        first.data_tables[0].contents,
        _records()[2].data_tables[0].contents,
    })
    workbook = load_workbook(master_path, read_only=True)
    assert workbook["Statistics"]["AA2"].value == "0.01234567890123456789"
    workbook.close()

    public_path = tmp_path / "public.xlsx"
    public = build_publication_workbook(
        folder,
        public_path,
        profile="public",
        experiment_statistics=ledger,
        safe_columns=["group", "value"],
    )
    assert public.valid
    package = public_path.read_bytes()
    assert b"participant-001" not in package

    minimal_path = tmp_path / "minimal.xlsx"
    minimal = build_publication_workbook(
        folder,
        minimal_path,
        profile="minimal-public",
        experiment_statistics=ledger,
        safe_columns=["group", "value"],
    )
    assert minimal.valid
    workbook = load_workbook(minimal_path, read_only=True)
    assert workbook.sheetnames == [
        "README", "Figures", "Data_Index", "Statistics", "Test_Families",
        "Verification", "Dictionary",
    ]
    workbook.close()


def test_signed_protected_workbook_and_failed_policy_are_atomic(tmp_path):
    folder, first, _second, _third = _mixed_artifacts(tmp_path)
    ledger = _ledger(tmp_path)
    key = tmp_path / "signing.pem"
    generate_signing_key(key, password="signing password")
    fingerprint = public_key_fingerprint(signing_public_bytes(
        load_signing_private_key(key, password="signing password")
    ))
    trusted_path = tmp_path / "trust.json"
    TrustStore([TrustEntry(fingerprint, "journal signer", scopes=["figure"])]).save(trusted_path)

    table_id = f"table:{first.data_tables[0].sha256}"
    output = tmp_path / "signed-protected.xlsx"
    result = build_publication_workbook(
        folder,
        output,
        experiment_statistics=ledger,
        protected_sections=[table_id],
        encryption_password="data-review password",
        signing_key_path=key,
        signing_password="signing password",
        required_meanings=["signature_valid", "signer_trusted"],
        trust_store=trusted_path,
    )
    assert result.valid
    assert result.proof["meanings"]["signature_valid"] == "pass"
    assert result.proof["meanings"]["signer_trusted"] == "pass"

    empty_trust = tmp_path / "empty-trust.json"
    TrustStore().save(empty_trust)
    untouched = tmp_path / "untouched.xlsx"
    sentinel = b"existing publication artifact"
    untouched.write_bytes(sentinel)
    with pytest.raises(ValueError, match="failed required proof policy"):
        build_publication_workbook(
            folder,
            untouched,
            experiment_statistics=ledger,
            signing_key_path=key,
            signing_password="signing password",
            required_meanings=["signer_trusted"],
            trust_store=empty_trust,
            overwrite=True,
        )
    assert untouched.read_bytes() == sentinel
